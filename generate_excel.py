#!/usr/bin/env python3
import sys, json, base64
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

def hex_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def border_all(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color="1B3A6B", size="medium"):
    return Border(bottom=Side(style=size, color=color))

AZUL        = "1B3A6B"
AZUL_MED    = "2563EB"
AZUL_CLARO  = "DBEAFE"
AZUL_HEADER = "1E40AF"
BLANCO      = "FFFFFF"
GRIS        = "F8FAFC"
GRIS_MED    = "E2E8F0"
VERDE       = "065F46"
VERDE_BG    = "D1FAE5"
ROJO        = "991B1B"
ROJO_BG     = "FEE2E2"
AMARILLO_BG = "FEF3C7"
AMARILLO    = "92400E"

def header_cell(ws, row, col, text, size=11, bold=True, color=BLANCO, bg=AZUL, wrap=False, center=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", size=size, bold=bold, color=color)
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=wrap)
    return c

def data_cell(ws, row, col, value, bold=False, color="1E293B", bg=BLANCO,
              fmt=None, center=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold, color=color, italic=italic)
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=True)
    if fmt: c.number_format = fmt
    return c

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def section_title(ws, row, col, text, span_end, bg=AZUL_HEADER):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", size=12, bold=True, color=BLANCO)
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span_end)
    ws.row_dimensions[row].height = 22

def generate_excel(data_json, output_path):
    data = json.loads(data_json)
    actual  = data["semana_actual"]
    anterior= data["semana_anterior"]
    ordenes = data.get("ordenes", [])
    empresa = data.get("empresa", "Manufacturas del Norte SA")
    reporte_md = data.get("reporte", "")

    wb = Workbook()

    # ══════════════════════════════════════════════
    # HOJA 1 — RESUMEN EJECUTIVO
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_view.showGridLines = False
    set_col_widths(ws1, [2, 22, 16, 16, 16, 16, 2])

    # Banner superior
    ws1.row_dimensions[1].height = 8
    ws1.row_dimensions[2].height = 50
    for col in range(2, 7):
        ws1.cell(row=2, column=col).fill = hex_fill(AZUL)
    ws1.merge_cells("B2:F2")
    c = ws1.cell(row=2, column=2)
    c.value = f"REPORTE EJECUTIVO DE PRODUCCIÓN"
    c.font = Font(name="Arial", size=18, bold=True, color=BLANCO)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws1.row_dimensions[3].height = 28
    for col in range(2, 7):
        ws1.cell(row=3, column=col).fill = hex_fill(AZUL_MED)
    ws1.merge_cells("B3:F3")
    c = ws1.cell(row=3, column=2)
    c.value = f"{empresa}   ·   {actual['periodo']}   ·   Generado: {datetime.date.today().strftime('%d/%m/%Y')}"
    c.font = Font(name="Arial", size=10, color=BLANCO)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws1.row_dimensions[4].height = 12

    # KPI Cards row
    ws1.row_dimensions[5].height = 20
    ws1.row_dimensions[6].height = 40
    ws1.row_dimensions[7].height = 25
    ws1.row_dimensions[8].height = 12

    turnos = actual["turnos"]
    total_prod = sum(t["piezas_ok"] for t in turnos)
    total_prog = sum(t["piezas_prog"] for t in turnos)
    total_rech = sum(t["rechazos"] for t in turnos)
    total_paro = sum(t["paro_min"] for t in turnos)
    cumplimiento = total_prod / total_prog if total_prog else 0

    total_prod_ant = sum(t["piezas_ok"] for t in anterior["turnos"])
    total_rech_ant = sum(t["rechazos"] for t in anterior["turnos"])
    total_paro_ant = sum(t["paro_min"] for t in anterior["turnos"])

    kpis = [
        ("PIEZAS PRODUCIDAS", total_prod, f"{total_prod:,}", "vs {}: {:+,}".format(anterior["periodo"], total_prod - total_prod_ant), total_prod >= total_prod_ant),
        ("CUMPLIMIENTO", cumplimiento, f"{cumplimiento:.1%}", f"Meta: 95%", cumplimiento >= 0.95),
        ("RECHAZOS", total_rech, f"{total_rech:,}", f"vs ant: {total_rech - total_rech_ant:+,}", total_rech <= total_rech_ant),
        ("MINUTOS PARO", total_paro, f"{total_paro}", f"vs ant: {total_paro - total_paro_ant:+}", total_paro <= total_paro_ant),
    ]

    kpi_cols = [2, 3, 4, 5]
    for i, (label, val, display, sub, is_good) in enumerate(kpis):
        col = kpi_cols[i]
        bg = VERDE_BG if is_good else ROJO_BG
        txt = VERDE if is_good else ROJO

        ws1.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col)
        lc = ws1.cell(row=5, column=col, value=label)
        lc.font = Font(name="Arial", size=8, bold=True, color="64748B")
        lc.fill = hex_fill(GRIS_MED)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        vc = ws1.cell(row=6, column=col, value=display)
        vc.font = Font(name="Arial", size=20, bold=True, color=txt)
        vc.fill = hex_fill(bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")

        sc = ws1.cell(row=7, column=col, value=sub)
        sc.font = Font(name="Arial", size=8, color="64748B", italic=True)
        sc.fill = hex_fill(bg)
        sc.alignment = Alignment(horizontal="center", vertical="center")

        for r in [5, 6, 7]:
            ws1.cell(row=r, column=col).border = border_all("CBD5E1")

    # Inventario alertas
    ws1.row_dimensions[9].height = 12
    section_title(ws1, 10, 2, "  ⚠  ALERTAS DE INVENTARIO", 5)
    ws1.row_dimensions[11].height = 18

    inv_headers = ["Material", "Stock Actual", "Mínimo", "Estado", "Acción"]
    for ci, h in enumerate(inv_headers, 2):
        c = ws1.cell(row=11, column=ci, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color=BLANCO)
        c.fill = hex_fill(AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all()

    for ri, mat in enumerate(actual["inventario"], 12):
        ws1.row_dimensions[ri].height = 16
        es_bajo = mat["stock"] < mat["minimo"]
        bg = ROJO_BG if es_bajo else VERDE_BG
        estado = "⚠ BAJO MÍNIMO" if es_bajo else "✓ OK"
        estado_color = ROJO if es_bajo else VERDE
        accion = f"Ordenar {mat['minimo']*4 - mat['stock']:,} {mat['unidad']}" if es_bajo else "—"

        vals = [mat["material"], f"{mat['stock']:,} {mat['unidad']}", f"{mat['minimo']:,} {mat['unidad']}", estado, accion]
        for ci, v in enumerate(vals, 2):
            c = ws1.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=9,
                          bold=(ci == 5 and es_bajo),
                          color=estado_color if ci == 4 else "1E293B")
            c.fill = hex_fill(bg if ci >= 4 else BLANCO)
            c.alignment = Alignment(horizontal="center" if ci >= 3 else "left", vertical="center")
            c.border = border_all()

    set_col_widths(ws1, [2, 28, 14, 14, 16, 28, 2])

    # ══════════════════════════════════════════════
    # HOJA 2 — PRODUCCIÓN POR TURNO
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("Producción por Turno")
    ws2.sheet_view.showGridLines = False
    set_col_widths(ws2, [2, 22, 14, 14, 10, 10, 10, 22, 2])

    ws2.row_dimensions[2].height = 36
    ws2.merge_cells("B2:H2")
    c = ws2.cell(row=2, column=2, value="PRODUCCIÓN POR TURNO — " + actual["periodo"])
    c.font = Font(name="Arial", size=14, bold=True, color=BLANCO)
    c.fill = hex_fill(AZUL)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws2.row_dimensions[4].height = 18
    headers = ["Turno", "Prog.", "Producidas", "Rechazos", "% Rechazo", "Paro (min)", "% Cumpl.", "Causa de Paro"]
    for ci, h in enumerate(headers, 2):
        c = ws2.cell(row=4, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=BLANCO)
        c.fill = hex_fill(AZUL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all()

    turno_colors = [("2563EB", "DBEAFE"), ("059669", "D1FAE5"), ("D97706", "FEF3C7")]

    for ri, (t, (accent, bg)) in enumerate(zip(turnos, turno_colors), 5):
        ws2.row_dimensions[ri].height = 18
        cumpl = t["piezas_ok"] / t["piezas_prog"] if t["piezas_prog"] else 0
        pct_rech = t["rechazos"] / t["piezas_ok"] if t["piezas_ok"] else 0
        is_bad = cumpl < 0.92

        row_bg = ROJO_BG if is_bad else bg
        vals = [t["turno"], t["piezas_prog"], t["piezas_ok"], t["rechazos"],
                pct_rech, t["paro_min"], cumpl, t["causa"]]
        fmts = [None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "0.0%", None]

        for ci, (v, fmt) in enumerate(zip(vals, fmts), 2):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10,
                          bold=(ci == 2),
                          color=accent if ci == 2 else "1E293B")
            c.fill = hex_fill(row_bg)
            c.alignment = Alignment(horizontal="center" if ci != 9 else "left",
                                     vertical="center", wrap_text=(ci == 9))
            if fmt: c.number_format = fmt
            c.border = border_all()

    # Fila totales
    ri_tot = 5 + len(turnos)
    ws2.row_dimensions[ri_tot].height = 20
    tot_vals = ["TOTAL SEMANA", f"=SUM(C5:C{ri_tot-1})", f"=SUM(D5:D{ri_tot-1})",
                f"=SUM(E5:E{ri_tot-1})", f"=E{ri_tot}/D{ri_tot}", f"=SUM(G5:G{ri_tot-1})",
                f"=D{ri_tot}/C{ri_tot}", "—"]
    tot_fmts = [None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "0.0%", None]
    for ci, (v, fmt) in enumerate(zip(tot_vals, tot_fmts), 2):
        c = ws2.cell(row=ri_tot, column=ci, value=v)
        c.font = Font(name="Arial", size=10, bold=True, color=BLANCO)
        c.fill = hex_fill(AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt: c.number_format = fmt
        c.border = border_all(BLANCO)

    # ══════════════════════════════════════════════
    # HOJA 3 — COMPARATIVO
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("Comparativo Semanal")
    ws3.sheet_view.showGridLines = False
    set_col_widths(ws3, [2, 24, 16, 16, 16, 16, 2])

    ws3.row_dimensions[2].height = 36
    ws3.merge_cells("B2:F2")
    c = ws3.cell(row=2, column=2, value=f"COMPARATIVO: {actual['periodo']} vs {anterior['periodo']}")
    c.font = Font(name="Arial", size=13, bold=True, color=BLANCO)
    c.fill = hex_fill(AZUL)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws3.row_dimensions[4].height = 18
    for ci, h in enumerate(["Indicador", "Semana Actual", "Semana Anterior", "Variación", "Tendencia"], 2):
        c = ws3.cell(row=4, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=BLANCO)
        c.fill = hex_fill(AZUL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all()

    metricas = [
        ("Piezas Producidas",  total_prod, total_prod_ant, True,  "#,##0"),
        ("Piezas Rechazadas",  total_rech, total_rech_ant, False, "#,##0"),
        ("Minutos de Paro",    total_paro, total_paro_ant, False, "#,##0"),
        ("% Cumplimiento",     cumplimiento, total_prod_ant/sum(t["piezas_prog"] for t in anterior["turnos"]) if sum(t["piezas_prog"] for t in anterior["turnos"]) else 0, True, "0.0%"),
        ("Tasa de Rechazo",    total_rech/total_prod if total_prod else 0, total_rech_ant/total_prod_ant if total_prod_ant else 0, False, "0.0%"),
    ]

    for ri, (label, act_v, ant_v, higher_is_better, fmt) in enumerate(metricas, 5):
        ws3.row_dimensions[ri].height = 18
        variacion = act_v - ant_v
        is_good = (variacion > 0) if higher_is_better else (variacion < 0)
        tendencia = "▲ Mejora" if is_good else ("▼ Deterioro" if variacion != 0 else "→ Igual")
        bg = VERDE_BG if is_good else ROJO_BG
        txt_color = VERDE if is_good else ROJO

        for ci, (v, f) in enumerate([(label, None), (act_v, fmt), (ant_v, fmt), (variacion, fmt), (tendencia, None)], 2):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10,
                          bold=(ci == 2),
                          color=txt_color if ci >= 4 else "1E293B")
            c.fill = hex_fill(bg if ci >= 4 else (GRIS if ri % 2 == 0 else BLANCO))
            c.alignment = Alignment(horizontal="left" if ci == 2 else "center", vertical="center")
            if f: c.number_format = f
            c.border = border_all()

    # ══════════════════════════════════════════════
    # HOJA 4 — ÓRDENES DE COMPRA
    # ══════════════════════════════════════════════
    ws4 = wb.create_sheet("Órdenes de Compra")
    ws4.sheet_view.showGridLines = False
    set_col_widths(ws4, [2, 14, 22, 14, 10, 12, 14, 12, 22, 2])

    ws4.row_dimensions[2].height = 36
    ws4.merge_cells("B2:I2")
    c = ws4.cell(row=2, column=2, value="ÓRDENES DE COMPRA GENERADAS POR AGENTE IA")
    c.font = Font(name="Arial", size=13, bold=True, color=BLANCO)
    c.fill = hex_fill(AZUL)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws4.row_dimensions[3].height = 20
    ws4.merge_cells("B3:I3")
    c = ws4.cell(row=3, column=2, value=f"Generadas automáticamente · {datetime.date.today().strftime('%d/%m/%Y')} · {empresa}")
    c.font = Font(name="Arial", size=9, italic=True, color="64748B")
    c.fill = hex_fill(AZUL_CLARO)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws4.row_dimensions[5].height = 18
    oc_headers = ["Folio", "Material", "Cantidad", "Unidad", "Precio Unit.", "Total MXN", "Entrega", "Proveedor", "Justificación"]
    for ci, h in enumerate(oc_headers, 2):
        c = ws4.cell(row=5, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=BLANCO)
        c.fill = hex_fill(AZUL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all()

    total_general = 0
    for ri, o in enumerate(ordenes, 6):
        ws4.row_dimensions[ri].height = 18
        bg = GRIS if ri % 2 == 0 else BLANCO
        total_mxn = float(o.get("total_mxn", 0))
        total_general += total_mxn
        vals = [o.get("folio",""), o.get("material",""), o.get("cantidad",0),
                o.get("unidad",""), o.get("precio_unitario",0), total_mxn,
                o.get("tiempo_entrega",""), o.get("proveedor",""), o.get("justificacion","")]
        fmts = [None, None, "#,##0", None, '"$"#,##0.00', '"$"#,##0.00', None, None, None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 2):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10,
                          bold=(ci == 2),
                          color=AZUL_MED if ci == 2 else "1E293B")
            c.fill = hex_fill(bg)
            c.alignment = Alignment(horizontal="center" if ci not in [3, 9] else "left",
                                     vertical="center", wrap_text=(ci == 9))
            if fmt: c.number_format = fmt
            c.border = border_all()

    if ordenes:
        ri_tot = 6 + len(ordenes)
        ws4.row_dimensions[ri_tot].height = 20
        ws4.merge_cells(start_row=ri_tot, start_column=2, end_row=ri_tot, end_column=5)
        c = ws4.cell(row=ri_tot, column=2, value="TOTAL INVERSIÓN REQUERIDA")
        c.font = Font(name="Arial", size=11, bold=True, color=BLANCO)
        c.fill = hex_fill(AZUL)
        c.alignment = Alignment(horizontal="right", vertical="center")

        tc = ws4.cell(row=ri_tot, column=6, value=total_general)
        tc.font = Font(name="Arial", size=12, bold=True, color=BLANCO)
        tc.fill = hex_fill(AZUL)
        tc.number_format = '"$"#,##0.00'
        tc.alignment = Alignment(horizontal="center", vertical="center")

        for col in [7, 8, 9]:
            ws4.cell(row=ri_tot, column=col).fill = hex_fill(AZUL)

    wb.save(output_path)
    print(json.dumps({"status": "ok", "path": output_path}))

if __name__ == "__main__":
    data_json = base64.b64decode(sys.argv[1]).decode('utf-8')
    output_path = sys.argv[2]
    generate_excel(data_json, output_path)
